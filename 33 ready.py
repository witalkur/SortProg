from tkinter import *



def do_it():
    import openpyxl
    import math
    from openpyxl import Workbook
    wb = openpyxl.load_workbook('группировать.xlsx')
    sheet = wb.active
    ListOfKeys = []
    wb2 = Workbook()
    ws = wb2.active

    for cell in list(sheet.columns)[0]:
        if cell.value != None:
            ListOfKeys.append(cell.value)
    ListOfValues = [cell.value for cell in list(sheet.columns)[1]]
    dict1 = dict(zip(ListOfKeys, ListOfValues))

    zag = int(Zagotovka.get())
    zapilz = int(zapil.get())
    '''
    if zapilz > 100:
        child = Toplevel(win)
        child.title('Ошибка')
        child.geometry('200x150')
        but2 = Button(child, text = 'ok', command = quit)
        but2.pack()
        '''
    zapasz = int(zapas.get())
    
    times = 0
    ListOfPares = []

    while times < 50:
            max_length = 0
            for TimberId, TimberLenght in dict1.items():
                if TimberLenght > max_length:
                    max_length = TimberLenght
                    max_timber_id = TimberId
            TheRest = 300
            dict2 = { k : v for k,v in dict1.items()}
            for TimberId1, TimberLenght1 in dict2.items():
                    cur_rest = (zag - (int(max_length) + int(TimberLenght1) + zapilz + zapasz))
                    zagot = (int(max_length) + int(TimberLenght1) + zapilz + zapasz)
                    if cur_rest >= 0 and cur_rest < TheRest and TimberId1 != max_timber_id:
                            TheRest = cur_rest
                            Pare = (TimberId, TimberId1, zagot)
                            ListOfPares.append(Pare)
                            x = dict1.pop(TimberId1, None)
                            y = dict1.pop(max_timber_id, None)
                            break
            times += 1
    times3 = 0
    while times3 < 50:        
        dict2 = dict1.copy()
        dict3 = dict2.copy()
        TheRest = 300
        for TimberId, TimberLenght in dict2.items():
            if TimberId not in list(dict1.keys()):
                break
            else:
                for TimberIdtemp, TimberLenghtTemp in dict3.items():
                    cur_rest = (zag - (int(TimberLenght) + int(TimberLenghtTemp) + zapilz + zapasz))
                    zagot = (int(TimberLenght) + int(TimberLenghtTemp) + zapilz + zapasz)
                    if cur_rest >= 0 and cur_rest < TheRest and TimberId != TimberIdtemp:
                        TheRest = cur_rest
                        Pare = (TimberId, TimberIdtemp, zagot)
                        ListOfPares.append(Pare)
                        x = dict1.pop(TimberId, None)
                        y = dict1.pop(TimberIdtemp, None)
                        break
        times3 += 1



    times4 = 0
    while times4 < 20:
        do_loop = True
        dict4 = dict1.copy()
        TheRest = 300
        for TimberId, TimberLenght in dict4.items():
            if TimberId not in list(dict1.keys()) or not do_loop:
                break
            else:
                dict5 = dict1.copy()
                for TimberIdtemp, TimberLenghtTemp in dict5.items():
                    if TimberIdtemp not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or not do_loop:
                        break
                    else:
                        if TimberId == TimberIdtemp:
                            pass
                        else:
                            dict6 = dict1.copy()
                            for TimberIdtemp2, TimberLenghtTemp2 in dict6.items():
                                if TimberIdtemp not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or TimberIdtemp2 not in list(dict1.keys()) or not do_loop:
                                    break
                                elif TimberId == TimberIdtemp or TimberId == TimberIdtemp2 or TimberIdtemp == TimberIdtemp2:
                                    pass
                                else:
                                    cur_rest = (zag - (int(TimberLenght) + int(TimberLenghtTemp) + int(TimberLenghtTemp2) + (zapilz * 2) + zapasz))
                                    zagot = (int(TimberLenght) + int(TimberLenghtTemp) + int(TimberLenghtTemp2) + (zapilz * 2) + zapasz)
                                    if cur_rest >= 0 and cur_rest < TheRest and TimberLenght != None and TimberLenghtTemp != None and TimberLenghtTemp2 != None and TimberId != TimberIdtemp and TimberId != TimberIdtemp2 and TimberIdtemp != TimberIdtemp2:
                                        TheRest = cur_rest
                                        Pare = (TimberId, TimberIdtemp, TimberIdtemp2, zagot)
                                        ListOfPares.append(Pare)
                                        x = dict1.pop(TimberId, None)
                                        y = dict1.pop(TimberIdtemp, None)
                                        z = dict1.pop(TimberIdtemp2, None)
                                        do_loop = False
                                        break
                    
        times4 += 1

    times5 = 0
    while times5 < 20:
        do_loop = True
        dict7 = dict1.copy()
        TheRest = 300
        for TimberId, TimberLenght in dict7.items():
            if TimberId not in list(dict1.keys()) or not do_loop:
                break
            else:
                dict8 = dict1.copy()
                for TimberIdtemp, TimberLenghtTemp in dict8.items():
                    if TimberIdtemp not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or not do_loop:
                        break
                    else:
                        if TimberId == TimberIdtemp:
                            pass
                        else:
                            dict9 = dict1.copy()
                            for TimberIdtemp3, TimberLenghtTemp3 in dict9.items():
                                if TimberIdtemp3 not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or TimberIdtemp not in list(dict1.keys()) or not do_loop:
                                    break
                                else:
                                    if TimberId == TimberIdtemp or TimberId == TimberIdtemp3 or TimberIdtemp == TimberIdtemp3: 
                                        pass
                                    else:
                                        dict10 = dict1.copy()
                                        for TimberIdtemp2, TimberLenghtTemp2 in dict10.items():
                                            if TimberIdtemp not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or TimberIdtemp2 not in list(dict1.keys()) or TimberIdtemp3 not in list(dict1.keys()) or not do_loop:
                                                break
                                            elif TimberId == TimberIdtemp or TimberId == TimberIdtemp2 or TimberIdtemp == TimberIdtemp2 or TimberId == TimberIdtemp3 or TimberIdtemp3 == TimberIdtemp2:
                                                pass
                                            else:
                                                cur_rest = (zag - (int(TimberLenght) + int(TimberLenghtTemp) + int(TimberLenghtTemp2) + int(TimberLenghtTemp3) + (zapilz * 2) + zapasz))
                                                zagot = (int(TimberLenght) + int(TimberLenghtTemp) + int(TimberLenghtTemp2) + int(TimberLenghtTemp3) + (zapilz * 3) + zapasz)
                                                if cur_rest >= 0 and cur_rest < TheRest and TimberLenght != None and TimberLenghtTemp3 != None and TimberLenghtTemp != None and TimberLenghtTemp2 != None and TimberId != TimberIdtemp and TimberId != TimberIdtemp2 and TimberIdtemp != TimberIdtemp2:
                                                    TheRest = cur_rest
                                                    Pare = (TimberId, TimberIdtemp, TimberIdtemp2, TimberIdtemp3, zagot)
                                                    ListOfPares.append(Pare)
                                                    x = dict1.pop(TimberId, None)
                                                    y = dict1.pop(TimberIdtemp, None)
                                                    z = dict1.pop(TimberIdtemp2, None)
                                                    w = dict1.pop(TimberIdtemp3, None)
                                                    do_loop = False
                                                    break
                    
        times5 += 1


    times6 = 0
    while times6 < 20:
        do_loop = True
        dict11 = dict1.copy()
        TheRest = 300
        for TimberId, TimberLenght in dict11.items():
            if TimberId not in list(dict1.keys()) or not do_loop:
                break
            else:
                dict12 = dict1.copy()
                for TimberIdtemp, TimberLenghtTemp in dict12.items():
                    if TimberIdtemp not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or not do_loop:
                        break
                    else:
                        if TimberId == TimberIdtemp:
                            pass
                        else:
                            dict13 = dict1.copy()
                            for TimberIdtemp3, TimberLenghtTemp3 in dict13.items():
                                if TimberIdtemp3 not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or TimberIdtemp not in list(dict1.keys()) or not do_loop:
                                    break
                                else:
                                    if TimberId == TimberIdtemp or TimberId == TimberIdtemp3 or TimberIdtemp == TimberIdtemp3: 
                                        pass
                                    else:
                                        dict14 = dict1.copy()
                                        for TimberIdtemp4, TimberLenghtTemp4 in dict14.items():
                                            if TimberIdtemp3 not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or TimberIdtemp4 not in list(dict1.keys()) or TimberIdtemp not in list(dict1.keys()) or not do_loop:
                                                break
                                            else:
                                                if TimberId == TimberIdtemp or TimberId == TimberIdtemp3 or TimberIdtemp == TimberIdtemp3 or TimberId == TimberIdtemp4 or TimberIdtemp3 == TimberIdtemp4: 
                                                    pass
                                                else:
                                                    dict14 = dict1.copy()
                                                    for TimberIdtemp2, TimberLenghtTemp2 in dict14.items():
                                                        if TimberIdtemp not in list(dict1.keys()) or TimberId not in list(dict1.keys()) or TimberIdtemp4 not in list(dict1.keys()) or TimberIdtemp2 not in list(dict1.keys()) or TimberIdtemp3 not in list(dict1.keys()) or not do_loop:
                                                            break
                                                        elif TimberId == TimberIdtemp or TimberId == TimberIdtemp4 or TimberIdtemp2 == TimberIdtemp4 or TimberIdtemp3 == TimberIdtemp4 or TimberIdtemp == TimberIdtemp4 or TimberId == TimberIdtemp2 or TimberIdtemp == TimberIdtemp2 or TimberId == TimberIdtemp3 or TimberIdtemp3 == TimberIdtemp2:
                                                            pass
                                                        else:
                                                            cur_rest = (zag - (int(TimberLenght) + int(TimberLenghtTemp) + int(TimberLenghtTemp2) + int(TimberLenghtTemp3) + int(TimberLenghtTemp4) + (zapilz * 2) + zapasz))
                                                            zagot = (int(TimberLenght) + int(TimberLenghtTemp) + int(TimberLenghtTemp2) + int(TimberLenghtTemp3) + int(TimberLenghtTemp4) + (zapilz * 4) + zapasz)
                                                            if cur_rest >= 0 and cur_rest < TheRest and TimberLenght != None and TimberLenghtTemp4 != None and TimberLenghtTemp3 != None and TimberLenghtTemp != None and TimberLenghtTemp2 != None and TimberId != TimberIdtemp and TimberId != TimberIdtemp2 and TimberIdtemp != TimberIdtemp2:
                                                                TheRest = cur_rest
                                                                Pare = (TimberId, TimberIdtemp, TimberIdtemp2, TimberIdtemp3, TimberIdtemp4, zagot)
                                                                ListOfPares.append(Pare)
                                                                x = dict1.pop(TimberId, None)
                                                                y = dict1.pop(TimberIdtemp, None)
                                                                z = dict1.pop(TimberIdtemp2, None)
                                                                w = dict1.pop(TimberIdtemp3, None)
                                                                n = dict1.pop(TimberIdtemp4, None)
                                                                do_loop = False
                                                                break
          
        times6 += 1


    from collections import Counter
    templist = []
    for key, value in dict1.items():
        templist.append(value)
    sorted(templist, key=int)
    tempdict = Counter(templist)
        
 
    p = 4
    text = 'Сгрупированные в ' + str(zag) + ' мм:'
    ws.cell(row=1, column=1).value = text
    ws.cell(row=2, column=1).value = 'Номера брусьев'
    ws.cell(row=2, column=6).value = 'Длина заготовки'

    for rec in ListOfPares:
        if len(rec) == 3:
            ws.cell(row=p, column=1).value = rec[0]
            ws.cell(row=p, column=2).value = rec[1]
            ws.cell(row=p, column=6).value = rec[2]
        if len(rec) == 4:
            ws.cell(row=p, column=1).value = rec[0]
            ws.cell(row=p, column=2).value = rec[1]
            ws.cell(row=p, column=3).value = rec[2]
            ws.cell(row=p, column=6).value = rec[3]
        if len(rec) == 5:
            ws.cell(row=p, column=1).value = rec[0]
            ws.cell(row=p, column=2).value = rec[1]
            ws.cell(row=p, column=3).value = rec[2]
            ws.cell(row=p, column=4).value = rec[3]
            ws.cell(row=p, column=6).value = rec[4]
        if len(rec) == 6:
            ws.cell(row=p, column=1).value = rec[0]
            ws.cell(row=p, column=2).value = rec[1]
            ws.cell(row=p, column=3).value = rec[2]
            ws.cell(row=p, column=4).value = rec[3]
            ws.cell(row=p, column=5).value = rec[4]
            ws.cell(row=p, column=6).value = rec[5]  
        p += 1
        
    r = 4
    ws.cell(row=1, column=9).value = 'Несгрупированные брусья'
    ws.cell(row=2, column=9).value = 'Номера брусьев'
    ws.cell(row=2, column=10).value = 'Длина брусьев'
    for TimberId, TimberLenght in dict1.items():
        ws.cell(row=r, column=9).value = TimberId
        ws.cell(row=r, column=10).value = TimberLenght
        r += 1

    l = 5
    
    ws.cell(row=3, column=12).value = 'Длина'
    ws.cell(row=3, column=13).value = 'Кол-во'
    
    for key, value in tempdict.items():
        ws.cell(row=l, column=12).value = key
        ws.cell(row=l, column=13).value = value
        l += 1
    
    wb2.save('результаты.xlsx')
    print('saved')

win = Tk()
win.geometry('500x200')

tl = Label(win, text='Сгруппировать в длину, мм', fg='blue')
tl.config(font=('Verdana', 14))
tl.pack()

Zagotovka = Entry(win, width = 20, bg = 'violet')
Zagotovka.pack()

t2 = Label(win, text='На запил , мм', fg='blue')
t2.config(font=('Verdana', 14))
t2.pack()

zapil = Entry(win, width = 20, bg = 'violet')
zapil.pack()

t3 = Label(win, text='Запас по длине, мм', fg='blue')
t3.config(font=('Verdana', 14))
t3.pack()

zapas = Entry(win, width = 20, bg = 'violet')
zapas.pack()

but = Button(win, text = 'Сгруппировать', command = do_it)
but.pack()
win.mainloop()
